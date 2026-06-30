import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_WORKFLOW = Path("/Users/jiangchuanchen/Downloads/sql优化_main_execute_workflow.json")
FILLED_XLSX = Path("/Users/jiangchuanchen/Downloads/异常告警部门账号通知 (3).xlsx")
MX_OVERRIDE_TSV = Path("/Users/jiangchuanchen/.codex/attachments/3ddf32a8-93b0-41a5-9822-a0844788d2b1/pasted-text.txt")
OUTPUT_WORKFLOW = Path("/Users/jiangchuanchen/Documents/Codex/2026-06-29/dan/outputs/sql优化_main_execute_workflow_部门账号联系人通知版.json")
OUTPUT_HARDCODED_WORKFLOW = Path("/Users/jiangchuanchen/Documents/Codex/2026-06-29/dan/outputs/sql优化_main_execute_workflow_部门账号联系人通知写死版.json")
OUTPUT_MAPPING = Path("/Users/jiangchuanchen/Documents/Codex/2026-06-29/dan/outputs/异常告警部门账号通知_mapping.json")
CONTACT_UPDATE_FORM_URL = "https://docs.google.com/forms/d/15Fa_fnvkrS5AwP8Zw127vSJymdvgf_lljMdpjzwCRVE/viewform"
CONTACT_OVERRIDE_CSV_URL = "https://docs.google.com/spreadsheets/d/16cZaIr0YupjC2P65O6Y3YdtJmznfoOqS-gq-bk1TGQI/export?format=csv&gid=998319904"

EMAIL_OVERRIDES_BY_CONTACT = {
    "罗玉瑰": ["rosieluo@kn.group"],
}

USER_EMAIL_OVERRIDES = {
    "alyssali": "Alyssali@weidu.co",
}

ACCOUNT_OVERRIDES = {
    ("cn", "e_ds_fox_dev"): {
        "department": "贷后业务组",
        "contacts": ["杜艳华"],
        "emails": ["elsadu@weidu.co"],
    },
    ("mx", "e_ds_strategy"): {
        "department": "信贷策略中心",
        "contacts": ["吴奎", "邓保保"],
        "emails": ["kuiwu@kn.group", "enzodeng@kn.group"],
    },
}


def clean(value):
    return "" if value is None else str(value).strip()


def split_people(value):
    return [
        item.strip()
        for item in re.split(r"[、,，;；/\\|\s]+", clean(value))
        if item.strip() and item.strip() not in {"联系人", "邮箱", "nan", "NaN", "NULL", "-"}
    ]


def find_col(ws, header):
    for cell in ws[1]:
        if clean(cell.value) == header:
            return cell.column
    return None


def parse_mx_override():
    if not MX_OVERRIDE_TSV.exists():
        return None

    lines = [line for line in MX_OVERRIDE_TSV.read_text(encoding="utf-8").splitlines() if clean(line)]
    if not lines:
        return None

    rows = []
    for line in lines[1:]:
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        country, account, department, contacts, emails = [clean(value) for value in parts[:5]]
        if country != "墨西哥" or not account:
            continue
        rows.append({
            "account": account,
            "department": department,
            "contacts": split_people(contacts),
            "emails": split_people(emails),
        })

    sheet_map = OrderedDict()
    for row in rows:
        if not row["emails"]:
            continue
        entry = sheet_map.setdefault(row["account"], {
            "department": row["department"],
            "contacts": [],
            "emails": [],
        })
        for contact in row["contacts"]:
            if contact not in entry["contacts"]:
                entry["contacts"].append(contact)
        for email in row["emails"]:
            if email not in entry["emails"]:
                entry["emails"].append(email)
    return sheet_map


def build_account_mapping():
    wb = load_workbook(FILLED_XLSX, data_only=True)
    mapping = OrderedDict()
    for ws in wb.worksheets:
        sheet_key = ws.title.lower()
        account_col = find_col(ws, "账号") or 2
        dept_col = find_col(ws, "部门") or 3
        contact_col = find_col(ws, "联系人") or 4
        email_col = find_col(ws, "邮箱")
        sheet_map = OrderedDict()
        if not email_col:
            continue

        for row in range(2, ws.max_row + 1):
            account = clean(ws.cell(row=row, column=account_col).value)
            if not account:
                continue
            contacts = split_people(ws.cell(row=row, column=contact_col).value)
            emails = split_people(ws.cell(row=row, column=email_col).value)
            override_emails = []
            for contact in contacts:
                override_emails.extend(EMAIL_OVERRIDES_BY_CONTACT.get(contact, []))
            if override_emails:
                emails = override_emails
            if not emails:
                continue

            entry = sheet_map.setdefault(account, {
                "department": clean(ws.cell(row=row, column=dept_col).value),
                "contacts": [],
                "emails": [],
            })
            for contact in contacts:
                if contact not in entry["contacts"]:
                    entry["contacts"].append(contact)
            for email in emails:
                if email not in entry["emails"]:
                    entry["emails"].append(email)
        mapping[sheet_key] = sheet_map
    mx_override = parse_mx_override()
    if mx_override is not None:
        mapping["mx"] = mx_override
    for (sheet, account), override in ACCOUNT_OVERRIDES.items():
        mapping.setdefault(sheet, OrderedDict())[account] = override
    return mapping


def build_merge_notify_target_js(account_mapping):
    mapping_json = json.dumps(account_mapping, ensure_ascii=False, separators=(",", ":"))
    user_email_overrides_json = json.dumps(USER_EMAIL_OVERRIDES, ensure_ascii=False, separators=(",", ":"))
    return f"""const safeFirst = (nodeName) => {{
  try {{
    const result = $(nodeName).first();
    return result && result.json ? result.json : null;
  }} catch (error) {{
    return null;
  }}
}};
const safeAll = (nodeName) => {{
  try {{
    return $(nodeName).all().map((item) => item.json || {{}}).filter((item) => item && Object.keys(item).length > 0);
  }} catch (error) {{
    return [];
  }}
}};
const normalizeId = (value) => String(value || '').trim().toLowerCase().replace(/^u_/i, '');
const normalizeAccount = (value) => String(value || '').trim().replace(/^u_/i, '');
const normalizeEmail = (value) => String(value || '').trim().toLowerCase();
const uniq = (values) => Array.from(new Set(values.map((value) => String(value || '').trim()).filter(Boolean)));
const notifyNode = safeFirst('Notify Config') || {{}};
const webhookNode = safeFirst('Webhook') || {{}};
const webhookBody = webhookNode.body || {{}};
const upstreamBase = safeFirst('Merge Optimized Explain Result')
  || safeFirst('Skip Optimized Explain')
  || safeFirst('Parse AI Result')
  || {{}};
const base = {{ ...upstreamBase, ...notifyNode }};
const config = base.notifyConfig || notifyNode || {{}};
const targetUser = normalizeId(base.user);
const targetExecutor = normalizeId(base.executor);
const webhookUser = normalizeId(webhookBody.user);
const webhookExecutor = normalizeId(webhookBody.executor);
const targetIds = Array.from(new Set([targetUser, targetExecutor, webhookUser, webhookExecutor].filter(Boolean)));
const rows = $input.all().map((item) => item.json).filter((row) => row && Object.keys(row).length > 0);
const rowIdentityValues = (row) => uniq([
  row.userId,
  row.user,
  String(row.user || '').replace(/^u_/i, ''),
  String(row.email || '').split('@')[0],
].map(normalizeId));
const matchedRows = rows.filter((row) => {{
  const rowEmail = String(row.email || '').trim();
  if (!rowEmail || !/@/.test(rowEmail)) return false;
  return rowIdentityValues(row).some((value) => targetIds.includes(value));
}});
const fallbackEmail = config.fallbackEmail || 'jiangchuanchen@kn.group';
const forceTestEmail = !!config.forceTestEmail;
const userEmailOverrides = {user_email_overrides_json};
const pickMatch = (...ids) => {{
  for (const id of ids) {{
    if (!id) continue;
    const matched = matchedRows.find((row) => rowIdentityValues(row).includes(id));
    if (matched && matched.email) return matched;
  }}
  return null;
}};
let primaryEmail = fallbackEmail;
if (!forceTestEmail) {{
  const matched = pickMatch(targetUser, webhookUser, targetExecutor, webhookExecutor);
  const overrideEmail = userEmailOverrides[targetUser] || userEmailOverrides[webhookUser] || userEmailOverrides[targetExecutor] || userEmailOverrides[webhookExecutor] || '';
  primaryEmail = String(overrideEmail || (matched && matched.email) || fallbackEmail).trim() || fallbackEmail;
}}
const accountNotifyMap = {mapping_json};
const splitPeople = (value) => String(value || '')
  .split(/[、,，;；/\\\\|\\n\\r\\t]+/)
  .map((item) => item.trim())
  .filter(Boolean);
const canonicalSheet = (value) => {{
  const text = String(value || '').trim().toLowerCase();
  if (!text) return '';
  if (/中国|china|starrocks_cn|\\bcn\\b/.test(text)) return 'cn';
  if (/印尼|indonesia|starrocks_ine|starrocks_id|\\bine\\b|\\bid\\b/.test(text)) return 'ine';
  if (/泰国|thailand|starrocks_th|\\bth\\b/.test(text)) return 'th';
  if (/菲律宾|philippines|starrocks_ph|\\bph\\b/.test(text)) return 'ph';
  if (/巴基斯坦|pakistan|starrocks_pak|starrocks_pk|\\bpak\\b|\\bpk\\b/.test(text)) return 'pk';
  if (/墨西哥|mexico|starrocks_mex|starrocks_mx|\\bmex\\b|\\bmx\\b/.test(text)) return 'mx';
  return text;
}};
const firstField = (row, names) => {{
  for (const name of names) {{
    if (row[name] !== undefined && row[name] !== null && String(row[name]).trim() !== '') return row[name];
  }}
  return '';
}};
const parseCsvRows = (text) => {{
  const rows = [];
  let row = [];
  let field = '';
  let inQuotes = false;
  const input = String(text || '');
  for (let i = 0; i < input.length; i += 1) {{
    const ch = input[i];
    if (ch === '"') {{
      if (inQuotes && input[i + 1] === '"') {{
        field += '"';
        i += 1;
      }} else {{
        inQuotes = !inQuotes;
      }}
    }} else if (ch === ',' && !inQuotes) {{
      row.push(field);
      field = '';
    }} else if ((ch === '\\n' || ch === '\\r') && !inQuotes) {{
      if (ch === '\\r' && input[i + 1] === '\\n') i += 1;
      row.push(field);
      if (row.some((cell) => String(cell).trim())) rows.push(row);
      row = [];
      field = '';
    }} else {{
      field += ch;
    }}
  }}
  row.push(field);
  if (row.some((cell) => String(cell).trim())) rows.push(row);
  if (rows.length < 2) return [];
  const headers = rows[0].map((cell) => String(cell || '').trim());
  return rows.slice(1).map((cells) => Object.fromEntries(headers.map((header, index) => [header, cells[index] || ''])));
}};
const normalizeOverrideRows = (rawItems) => {{
  const normalized = [];
  for (const item of rawItems) {{
    if (Array.isArray(item)) {{
      normalized.push(...normalizeOverrideRows(item));
      continue;
    }}
    const body = item.data || item.body || item.response || item;
    if (typeof body === 'string') {{
      const text = body.trim();
      if (!text) continue;
      try {{
        const parsed = JSON.parse(text);
        normalized.push(...normalizeOverrideRows(Array.isArray(parsed) ? parsed : [parsed]));
      }} catch (error) {{
        normalized.push(...parseCsvRows(text));
      }}
      continue;
    }}
    if (body && typeof body === 'object') {{
      if (Array.isArray(body)) normalized.push(...normalizeOverrideRows(body));
      else if (Array.isArray(body.rows)) normalized.push(...normalizeOverrideRows(body.rows));
      else if (Array.isArray(body.data)) normalized.push(...normalizeOverrideRows(body.data));
      else normalized.push(body);
    }}
  }}
  return normalized;
}};
const overrideCountryOwners = {{
  cn: {{ contacts: ['张恒阳', '宗美晨'], emails: ['owenzhang@kn.group', 'rockyzong@kn.group'] }},
  ine: {{ contacts: ['何柳琴', '翟爽'], emails: ['gretchenhe@kn.group', 'riverzhai@kn.group'] }},
  th: {{ contacts: ['黄启龙', '朱会明'], emails: ['qilonghuang@kn.group', 'brightonzhu@kn.group'] }},
  ph: {{ contacts: ['汤伟', '陈江川'], emails: ['simontang@kn.group', 'jiangchuanchen@kn.group'] }},
  pk: {{ contacts: ['余红叶', '穆晋杰'], emails: ['adamyu@kn.group', 'moonmu@kn.group'] }},
  mx: {{ contacts: ['吴奎', '邓保保'], emails: ['kuiwu@kn.group', 'enzodeng@kn.group'] }},
}};
const applyContactOverrides = (rows) => {{
  const applied = [];
  for (const row of rows) {{
    const sheet = canonicalSheet(firstField(row, ['国家', 'country', 'sheet', '集群', 'cluster']));
    const account = normalizeAccount(firstField(row, ['账号', 'account', 'user', '用户']));
    const department = String(firstField(row, ['部门', '数据开发组', 'department', 'dept']) || '').trim();
    const note = String(firstField(row, ['备注', '备注(近2月访问)', 'note', 'remark', 'remarks']) || '').trim();
    let contacts = splitPeople(firstField(row, ['修改后联系人', '联系人', 'contacts', 'contact', 'name', 'names']));
    let emails = splitPeople(firstField(row, ['修改后邮箱', '邮箱', 'emails', 'email']));
    const action = String(firstField(row, ['操作', 'action', '是否删除', 'delete']) || '').trim().toLowerCase();
    if (!sheet || !account) continue;
    if (action === 'delete' || action === '删除' || action === 'true') {{
      if (accountNotifyMap[sheet]) delete accountNotifyMap[sheet][account];
      applied.push({{ sheet, account, action: 'delete' }});
      continue;
    }}
    let defaultedToCountryOwner = false;
    if (!contacts.length && overrideCountryOwners[sheet]) {{
      contacts = overrideCountryOwners[sheet].contacts;
      emails = overrideCountryOwners[sheet].emails;
      defaultedToCountryOwner = true;
    }}
    if (/未填写.*联系人|兜底.*国家负责人|国家负责人.*确认/.test(note)) {{
      defaultedToCountryOwner = true;
    }}
    if (!emails.length) continue;
    accountNotifyMap[sheet] = accountNotifyMap[sheet] || {{}};
    accountNotifyMap[sheet][account] = {{ department, contacts, emails, defaultedToCountryOwner, note }};
    applied.push({{ sheet, account, department, contacts, emails, action: 'upsert', defaultedToCountryOwner, note }});
  }}
  return applied;
}};
const contactOverrideRawRows = normalizeOverrideRows(safeAll('Fetch Contact Overrides'));
const contactOverrideApplied = applyContactOverrides(contactOverrideRawRows);
const queryContext = (base.evidence || {{}}).queryContext || {{}};
const alert = (base.evidence || {{}}).alert || {{}};
const rawCluster = String(base.cluster || queryContext.cluster || webhookBody.cluster || '').toLowerCase();
const rawCountry = String(queryContext.country || base.country || webhookBody.country || '').toLowerCase();
const sheetAliases = {{
  cn: ['cn'], china: ['cn'], starrocks_cn: ['cn'],
  ine: ['ine'], id: ['ine'], indonesia: ['ine'], starrocks_ine: ['ine'], starrocks_id: ['ine'],
  th: ['th'], thailand: ['th'], starrocks_th: ['th'],
  ph: ['ph'], philippines: ['ph'], starrocks_ph: ['ph'],
  pk: ['pk'], pak: ['pk'], pakistan: ['pk'], starrocks_pk: ['pk'], starrocks_pak: ['pk'],
  mx: ['mx'], mex: ['mx'], mexico: ['mx'], starrocks_mx: ['mx'], starrocks_mex: ['mx']
}};
const inferSheets = (...keys) => {{
  const sheets = [];
  for (const key of keys) {{
    const normalized = String(key || '').toLowerCase().trim();
    if (!normalized) continue;
    for (const [alias, mappedSheets] of Object.entries(sheetAliases)) {{
      if (normalized === alias || normalized.includes(alias)) {{
        for (const sheet of mappedSheets) sheets.push(sheet);
      }}
    }}
  }}
  return uniq(sheets);
}};
const candidateSheets = inferSheets(rawCluster, rawCountry);
const accountCandidates = uniq([
  base.account,
  base.dbUser,
  base.user,
  base.executor,
  webhookBody.account,
  webhookBody.dbUser,
  webhookBody.user,
  webhookBody.executor,
  alert.account,
  alert.user,
  alert.executor,
  queryContext.account,
  queryContext.user,
  queryContext.executor,
].map(normalizeAccount));
const findAccountOwners = () => {{
  const searchSheets = candidateSheets.length ? candidateSheets : Object.keys(accountNotifyMap);
  const owners = [];
  for (const sheet of searchSheets) {{
    const sheetMap = accountNotifyMap[sheet] || {{}};
    for (const account of accountCandidates) {{
      const entry = sheetMap[account];
      if (!entry) continue;
      owners.push({{ sheet, account, ...entry }});
    }}
  }}
  return owners;
}};
const matchedAccountOwners = findAccountOwners();
const mappedEmails = uniq(matchedAccountOwners.flatMap((entry) => entry.emails || []));
const mappedContacts = uniq(matchedAccountOwners.flatMap((entry) => entry.contacts || []));
const mappedDepartments = uniq(matchedAccountOwners.map((entry) => entry.department));
const mappedAccounts = uniq(matchedAccountOwners.map((entry) => entry.account));
const countryOwnerFallbackAccounts = matchedAccountOwners
  .filter((entry) => entry.defaultedToCountryOwner)
  .map((entry) => entry.account);
const countryOwnerFallbackNotice = countryOwnerFallbackAccounts.length
  ? '该账号联系人由国家负责人兜底；请国家负责人自行确认实际负责人并通过联系人调整表单补充。'
  : '';
const finalEmails = forceTestEmail
  ? [fallbackEmail]
  : (mappedEmails.length ? mappedEmails : [primaryEmail]);
const ownerMap = {{
  starrocks_cn: [{{ name: '张恒阳', email: 'owenzhang@kn.group' }}, {{ name: '宗美晨', email: 'rockyzong@kn.group' }}],
  cn: [{{ name: '张恒阳', email: 'owenzhang@kn.group' }}, {{ name: '宗美晨', email: 'rockyzong@kn.group' }}],
  starrocks_ine: [{{ name: '何柳琴', email: 'gretchenhe@kn.group' }}, {{ name: '翟爽', email: 'riverzhai@kn.group' }}],
  ine: [{{ name: '何柳琴', email: 'gretchenhe@kn.group' }}, {{ name: '翟爽', email: 'riverzhai@kn.group' }}],
  id: [{{ name: '何柳琴', email: 'gretchenhe@kn.group' }}, {{ name: '翟爽', email: 'riverzhai@kn.group' }}],
  starrocks_mex: [{{ name: '吴奎', email: 'kuiwu@kn.group' }}, {{ name: '邓保保', email: 'enzodeng@kn.group' }}],
  mex: [{{ name: '吴奎', email: 'kuiwu@kn.group' }}, {{ name: '邓保保', email: 'enzodeng@kn.group' }}],
  mx: [{{ name: '吴奎', email: 'kuiwu@kn.group' }}, {{ name: '邓保保', email: 'enzodeng@kn.group' }}],
  starrocks_ph: [{{ name: '汤伟', email: 'simontang@kn.group' }}, {{ name: '陈江川', email: 'jiangchuanchen@kn.group' }}],
  ph: [{{ name: '汤伟', email: 'simontang@kn.group' }}, {{ name: '陈江川', email: 'jiangchuanchen@kn.group' }}],
  starrocks_th: [{{ name: '黄启龙', email: 'qilonghuang@kn.group' }}, {{ name: '朱会明', email: 'brightonzhu@kn.group' }}],
  th: [{{ name: '黄启龙', email: 'qilonghuang@kn.group' }}, {{ name: '朱会明', email: 'brightonzhu@kn.group' }}],
  starrocks_pak: [{{ name: '余红叶', email: 'adamyu@kn.group' }}, {{ name: '穆晋杰', email: 'moonmu@kn.group' }}],
  pak: [{{ name: '余红叶', email: 'adamyu@kn.group' }}, {{ name: '穆晋杰', email: 'moonmu@kn.group' }}],
  pk: [{{ name: '余红叶', email: 'adamyu@kn.group' }}, {{ name: '穆晋杰', email: 'moonmu@kn.group' }}]
}};
const clusterKey = String(base.cluster || queryContext.cluster || '').toLowerCase();
const countryKey = String(queryContext.country || base.country || '').toLowerCase();
const owners = ownerMap[clusterKey] || ownerMap[countryKey] || [];
const countryOwnerNames = owners.map((item) => item.name).filter(Boolean);
const countryOwnerNamesText = countryOwnerNames.length ? countryOwnerNames.join('、') : '未配置';
const departmentAccountContactText = mappedContacts.length
  ? mappedContacts.map((name, index) => {{
      const email = mappedEmails[index] || '';
      return email ? name + '<' + email + '>' : name;
    }}).join('、')
  : '';
return [{{ json: {{
  ...base,
  notifyEmail: finalEmails[0] || fallbackEmail,
  notifyEmails: finalEmails,
  notifyTargetSource: mappedEmails.length ? 'department_account_mapping' : 'executor_or_fallback',
  userInfo: matchedRows,
  notifyConfig: config,
  departmentAccountNotifyMatched: Boolean(mappedEmails.length),
  departmentAccountNotifySheets: candidateSheets,
  departmentAccountNotifyAccounts: mappedAccounts,
  departmentAccountNotifyDepartments: mappedDepartments,
  departmentAccountNotifyContacts: mappedContacts,
  departmentAccountNotifyEmails: mappedEmails,
  departmentAccountContactText,
  countryOwnerFallbackAccounts,
  countryOwnerFallbackNotice,
  contactOverrideApplied,
  contactOverrideAppliedCount: contactOverrideApplied.length,
  countryOwnerEmails: [],
  countryOwnerNames,
  countryOwnerNamesText,
  tvMentionEmails: finalEmails,
  tvContactText: countryOwnerNamesText
}}}}];"""


BUILD_SIDECAR_PAYLOAD_JS = """const base = $('Build Langfuse Batch').first().json || {};
const notifyConfig = base.notifyConfig || {};
const evidence = (base.evidence || {}).alert || {};
const ai = base.aiResult || {};
const toText = (value) => value === null || value === undefined ? '' : String(value);
const sanitize = (value) => toText(value).replace(/undefined|null/g, '').trim();
const uniq = (values) => Array.from(new Set(values.map((value) => sanitize(value)).filter(Boolean)));
const cluster = sanitize(base.cluster);
const queryId = sanitize(base.queryId);
const notifyEmails = uniq(Array.isArray(base.notifyEmails) && base.notifyEmails.length ? base.notifyEmails : [base.notifyEmail || 'jiangchuanchen@kn.group']);
const contactText = sanitize(base.countryOwnerNamesText) || '未配置';
const contactUpdateFormUrl = sanitize(notifyConfig.contactUpdateFormUrl);
const countryOwnerFallbackNotice = sanitize(base.countryOwnerFallbackNotice);
const explainConclusion = !base.optimizedExplainChecked
  ? '未执行优化后 EXPLAIN，请不要直接使用优化后的 SQL，建议结合上述优化建议人工验证。'
  : (base.optimizedExplainOk
    ? (base.optimizedExplainPermissionDenied ? '优化后 EXPLAIN 因权限不足未实际执行，请结合上述优化建议人工验证，确认无风险后再尝试使用优化后的 SQL。' : '优化后 EXPLAIN 校验通过，可以尝试使用优化后的 SQL。')
    : '优化后 EXPLAIN 失败，请不要直接使用优化后的 SQL，建议结合上述优化建议人工验证。');
const buildSuggestionText = () => '异常查询：\\n'
  + '- 集群：' + cluster + '\\n'
  + '- 优化后sql的详细地址：http://sr-admin.kuainiujinke.com/share/sql/' + queryId + '\\n'
  + '- CPU时间：' + sanitize(evidence.cpuTime) + '\\n'
  + '- 内存使用：' + sanitize(evidence.memUsage) + '\\n'
  + '- 执行时间：' + sanitize(evidence.execTime) + '\\n'
  + '- 扫描行数：' + sanitize(evidence.scanRows) + '\\n'
  + '- 用户：' + sanitize(base.user) + '\\n\\n'
  + '告警原因：\\n'
  + '- ' + sanitize(evidence.violationDetail) + '\\n\\n'
  + 'AI结论：\\n'
  + sanitize(ai.oneLineConclusion) + '\\n\\n'
  + '核心问题：\\n'
  + sanitize(ai.coreIssue) + '\\n\\n'
  + '优化建议：\\n'
  + sanitize(ai.summary) + '\\n\\n'
  + '优化风险点：\\n'
  + sanitize(ai.risk) + '\\n\\n'
  + '需人工验证：\\n'
  + sanitize(ai.validationAdvice) + '\\n\\n'
  + '优化后 EXPLAIN 校验：\\n'
  + explainConclusion + '\\n\\n'
  + (countryOwnerFallbackNotice ? '联系人兜底说明：\\n' + countryOwnerFallbackNotice + '\\n\\n' : '')
  + '如果需要人工协助，请联系：\\n'
  + contactText
  + (contactUpdateFormUrl ? '\\n\\n联系人调整表单：\\n' + contactUpdateFormUrl : '');
const data = buildSuggestionText();
const payloads = notifyEmails.map((email) => {
  const isWeiduEmail = /@weidu\\.co$/i.test(email);
  const accountId = sanitize(isWeiduEmail ? notifyConfig.weiduSidecarAccountId : notifyConfig.sidecarAccountId);
  const token = sanitize(isWeiduEmail ? notifyConfig.weiduSidecarToken : notifyConfig.sidecarToken);
  const payload = { type: 'TEXT', email, data, accountId, token };
  return {
    email,
    channel: isWeiduEmail ? 'weidu' : 'default',
    payload,
    shouldSend: Boolean(email && accountId && token && data.trim())
  };
});
return payloads.map((entry) => ({
  json: {
    ...base,
    notifyEmail: entry.email,
    sidecarShouldSend: entry.shouldSend,
    sidecarChannel: entry.channel,
    sidecarPayload: entry.payload,
    sidecarPayloads: payloads.map((item) => item.payload),
    sidecarRecipientCount: payloads.length
  }
}));"""


DEDUPLICATE_JS = """const inputItems = $input.all();
const staticData = typeof $getWorkflowStaticData === 'function'
  ? $getWorkflowStaticData('global')
  : (typeof getWorkflowStaticData === 'function' ? getWorkflowStaticData('global') : {});
const now = Date.now();
const dedupeWindowMs = 2 * 60 * 60 * 1000;
if (!staticData.queryNotifyDedupe || typeof staticData.queryNotifyDedupe !== 'object') {
  staticData.queryNotifyDedupe = {};
}
const store = staticData.queryNotifyDedupe;
for (const [key, value] of Object.entries(store)) {
  const timestamp = Number(value && value.sentAt ? value.sentAt : value);
  if (!Number.isFinite(timestamp) || now - timestamp > dedupeWindowMs) {
    delete store[key];
  }
}
return inputItems.map((item) => {
  const base = (item && item.json) || {};
  const queryId = String(base.queryId || '').trim();
  const recipientEmail = String((base.sidecarPayload || {}).email || base.notifyEmail || '').trim().toLowerCase();
  let duplicateNotifySuppressed = false;
  let duplicateNotifyReason = '';
  const dedupeKey = queryId && recipientEmail ? queryId + '::' + recipientEmail : queryId;
  if (dedupeKey) {
    const record = store[dedupeKey];
    const sentAt = Number(record && record.sentAt ? record.sentAt : record);
    if (Number.isFinite(sentAt) && now - sentAt <= dedupeWindowMs) {
      duplicateNotifySuppressed = true;
      duplicateNotifyReason = 'duplicate queryId/email within 2h: ' + dedupeKey;
    } else if (base.sidecarShouldSend) {
      store[dedupeKey] = {
        sentAt: now,
        sessionId: String(base.sessionId || ''),
        email: recipientEmail,
      };
    }
  }
  return { json: {
    ...base,
    sidecarShouldSend: Boolean(base.sidecarShouldSend) && !duplicateNotifySuppressed,
    duplicateNotifySuppressed,
    duplicateNotifyReason,
    duplicateNotifyWindowMs: dedupeWindowMs,
    duplicateNotifyKey: dedupeKey,
  }};
});"""


GET_USER_INFO_QUERY = """select email, userId, user
from employee
where (
  lower(coalesce(userId, '')) in (
    lower('{{ $('Notify Config').first().json.user }}'),
    lower('{{ $('Notify Config').first().json.executor }}'),
    lower('{{ $('Webhook').item.json.body.user }}'),
    lower('{{ $('Webhook').item.json.body.executor }}'),
    lower('{{ String($('Notify Config').first().json.user || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Notify Config').first().json.executor || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Webhook').item.json.body.user || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Webhook').item.json.body.executor || '').replace(/^u_/i, '') }}')
  )
  or lower(coalesce(user, '')) in (
    lower('{{ $('Notify Config').first().json.user }}'),
    lower('{{ $('Notify Config').first().json.executor }}'),
    lower('{{ $('Webhook').item.json.body.user }}'),
    lower('{{ $('Webhook').item.json.body.executor }}'),
    lower(concat('u_', '{{ String($('Notify Config').first().json.user || '').replace(/^u_/i, '') }}')),
    lower(concat('u_', '{{ String($('Notify Config').first().json.executor || '').replace(/^u_/i, '') }}')),
    lower(concat('u_', '{{ String($('Webhook').item.json.body.user || '').replace(/^u_/i, '') }}')),
    lower(concat('u_', '{{ String($('Webhook').item.json.body.executor || '').replace(/^u_/i, '') }}'))
  )
  or lower(substring_index(coalesce(email, ''), '@', 1)) in (
    lower('{{ String($('Notify Config').first().json.user || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Notify Config').first().json.executor || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Webhook').item.json.body.user || '').replace(/^u_/i, '') }}'),
    lower('{{ String($('Webhook').item.json.body.executor || '').replace(/^u_/i, '') }}')
  )
)
and email is not null
and trim(email) <> ''"""


def replace_node_code(workflow, name, code):
    for node in workflow["nodes"]:
        if node.get("name") == name:
            node.setdefault("parameters", {})["jsCode"] = code
            return
    raise ValueError(f"node not found: {name}")


def replace_node_query(workflow, name, query):
    for node in workflow["nodes"]:
        if node.get("name") == name:
            node.setdefault("parameters", {})["query"] = query
            return
    raise ValueError(f"node not found: {name}")


def patch_notify_config(workflow):
    for node in workflow["nodes"]:
        if node.get("name") == "Notify Config":
            code = node.setdefault("parameters", {}).get("jsCode", "")
            if "contactOverrideUrl" not in code:
                code = code.replace(
                    "weiduSidecarToken: '__FILL_IN_N8N__'",
                    "weiduSidecarToken: '__FILL_IN_N8N__',\n"
                    "  // Optional: set this to a Google Form response Sheet CSV/JSON URL or an Apps Script JSON endpoint.\n"
                    f"  contactOverrideUrl: '{CONTACT_OVERRIDE_CSV_URL}',\n"
                    "  // User-facing form link shown in each alert message.\n"
                    f"  contactUpdateFormUrl: '{CONTACT_UPDATE_FORM_URL}'"
                )
            elif "contactUpdateFormUrl" not in code:
                code = code.replace(
                    "contactOverrideUrl: ''",
                    f"contactOverrideUrl: '{CONTACT_OVERRIDE_CSV_URL}',\n"
                    "  // User-facing form link shown in each alert message.\n"
                    f"  contactUpdateFormUrl: '{CONTACT_UPDATE_FORM_URL}'"
                )
            node["parameters"]["jsCode"] = code
            return
    raise ValueError("node not found: Notify Config")


def ensure_dynamic_override_nodes(workflow):
    node_names = {node.get("name") for node in workflow["nodes"]}
    if "Has Contact Override URL?" not in node_names:
        workflow["nodes"].append({
            "parameters": {
                "conditions": {
                    "boolean": [
                        {
                            "value1": "={{ Boolean($json.contactOverrideUrl || (($json.notifyConfig || {}).contactOverrideUrl)) }}",
                            "value2": True,
                        }
                    ]
                },
                "options": {},
            },
            "id": "7a1dfbda-7f8b-4bd5-b872-6d5e671aa301",
            "name": "Has Contact Override URL?",
            "type": "n8n-nodes-base.if",
            "typeVersion": 2,
            "position": [52080, 21296],
        })
    if "Fetch Contact Overrides" not in node_names:
        workflow["nodes"].append({
            "parameters": {
                "method": "GET",
                "url": "={{ $json.contactOverrideUrl || (($json.notifyConfig || {}).contactOverrideUrl) }}",
                "responseFormat": "text",
                "sendHeaders": False,
                "options": {
                    "timeout": 30000
                },
            },
            "id": "74264abc-dfb8-4550-b4a8-0608d24ec9f4",
            "name": "Fetch Contact Overrides",
            "type": "n8n-nodes-base.httpRequest",
            "typeVersion": 4.4,
            "position": [52320, 21184],
            "onError": "continueRegularOutput",
        })

    connections = workflow.setdefault("connections", {})
    connections["Notify Config"] = {
        "main": [[{"node": "Has Contact Override URL?", "type": "main", "index": 0}]]
    }
    connections["Has Contact Override URL?"] = {
        "main": [
            [{"node": "Fetch Contact Overrides", "type": "main", "index": 0}],
            [{"node": "Get User Info", "type": "main", "index": 0}],
        ]
    }
    connections["Fetch Contact Overrides"] = {
        "main": [[{"node": "Get User Info", "type": "main", "index": 0}]]
    }


def write_outputs(workflow, account_mapping, workflow_path, workflow_name, use_google_overrides):
    if use_google_overrides:
        patch_notify_config(workflow)
        ensure_dynamic_override_nodes(workflow)
    else:
        # Hardcoded version: keep the original Notify Config -> Get User Info path and do not
        # attach the experimental Google Sheet override nodes.
        workflow.setdefault("connections", {})["Notify Config"] = {
            "main": [[{"node": "Get User Info", "type": "main", "index": 0}]]
        }

    replace_node_code(workflow, "Merge Notify Target", build_merge_notify_target_js(account_mapping))
    replace_node_query(workflow, "Get User Info", GET_USER_INFO_QUERY)
    replace_node_code(workflow, "Build Sidecar Payload", BUILD_SIDECAR_PAYLOAD_JS)
    replace_node_code(workflow, "Deduplicate Notify By QueryId", DEDUPLICATE_JS)

    workflow["name"] = workflow_name
    workflow_path.parent.mkdir(parents=True, exist_ok=True)
    workflow_path.write_text(json.dumps(workflow, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    workflow = json.loads(SOURCE_WORKFLOW.read_text(encoding="utf-8"))
    account_mapping = build_account_mapping()

    write_outputs(
        workflow=json.loads(SOURCE_WORKFLOW.read_text(encoding="utf-8")),
        account_mapping=account_mapping,
        workflow_path=OUTPUT_WORKFLOW,
        workflow_name="sql优化_main_execute_workflow_部门账号联系人通知版",
        use_google_overrides=True,
    )
    write_outputs(
        workflow=workflow,
        account_mapping=account_mapping,
        workflow_path=OUTPUT_HARDCODED_WORKFLOW,
        workflow_name="sql优化_main_execute_workflow_部门账号联系人通知写死版",
        use_google_overrides=False,
    )
    OUTPUT_MAPPING.write_text(json.dumps(account_mapping, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    accounts = sum(len(sheet_map) for sheet_map in account_mapping.values())
    multi_recipient = sum(
        1
        for sheet_map in account_mapping.values()
        for entry in sheet_map.values()
        if len(entry.get("emails", [])) > 1
    )
    print(json.dumps({
        "output": str(OUTPUT_WORKFLOW),
        "hardcoded_output": str(OUTPUT_HARDCODED_WORKFLOW),
        "mapping": str(OUTPUT_MAPPING),
        "sheets": list(account_mapping.keys()),
        "accounts": accounts,
        "multi_recipient_accounts": multi_recipient,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
